#!/usr/bin/env python3
"""
TrustPoint Network Map render service (detailed).

POST /render (JSON): { "client": "...", "date": "YYYY-MM-DD" (optional),
  "logo_url": "https://.../logo.png" (optional),
  "devices": [ {"name","class","ip","gw","dns","dhcp","mac","os",
                "site" (optional N-central site name)}, ... ] }
-> 200 image/png  (branded, detailed logical network map)
GET /healthz -> 200 "ok"
Requires: graphviz (dot) + Pillow + Flask + DejaVu fonts (see Dockerfile).

Detailed mode: every device is drawn individually with name, IP, role, OS and
MAC, grouped into network segments (by /24 subnet) that show the segment
gateway, DNS and DHCP. Firewalls and domain controllers are highlighted.
VLAN data is not available from N-central and is layered in later from
Network Detective (Phase 2).
"""
import io, re, ipaddress, datetime, subprocess, os, urllib.request, base64, json
from collections import OrderedDict
from flask import Flask, request, send_file, jsonify
from PIL import Image, ImageDraw, ImageFont

app = Flask(__name__)

NAVY = (18, 49, 94)
CONF = ("CONFIDENTIAL  -  Property of TrustPoint IT Solutions Inc.  -  "
        "Contains proprietary network information; for authorized internal use only, do not distribute.")
FONT_BOLD = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"
PUBLIC_DNS = ("8.8", "8.4", "1.1", "4.4", "9.9", "208.67", "209.18")

ROLE_COLOR = {
    "firewall": ("#b3261e", "white"), "switch": ("#1c7293", "white"),
    "network": ("#1c7293", "white"), "accesspoint": ("#2c5f8a", "white"),
    "nas": ("#6d2e46", "white"), "hyperv": ("#2c5f2d", "white"),
    "server": ("#1E2761", "white"), "printer": ("#8d6e8f", "white"),
    "endpoint": ("#cfe0f5", "#1E2761"), "laptop": ("#e2ecf9", "#1E2761"),
    # Network Glue port-fingerprinted roles
    "oracle": ("#8a3b00", "white"), "sql": ("#7a1f2b", "white"),
    "dns": ("#155e63", "white"), "linux": ("#33475b", "white"),
    "web": ("#3a4a2f", "white"),
    "other": ("#b8b8b8", "#222222"),
}
ROLE_NAMES = {
    "firewall": "FIREWALL", "switch": "SWITCH", "network": "ROUTER / SWITCH",
    "accesspoint": "ACCESS POINT", "nas": "NAS / STORAGE", "hyperv": "HYPER-V HOST",
    "server": "SERVER", "printer": "PRINTER", "endpoint": "WORKSTATION",
    "laptop": "LAPTOP", "oracle": "ORACLE DB", "sql": "SQL SERVER",
    "dns": "DNS / GATEWAY", "linux": "LINUX / SSH HOST", "web": "WEB / HTTPS",
    "other": "DEVICE",
}

# Port -> role (Network Glue "listening-ports"). First match wins.
PORT_ROLE = [(1521, "oracle"), (1433, "sql"), (53, "dns"),
             (9100, "printer"), (515, "printer"), (631, "printer"),
             (22, "linux"), (443, "web"), (8080, "web"), (80, "web")]

def parse_ports(v):
    """Accept [22,80] or 'SSH (22/TCP),HTTP (8080/TCP)' -> set of ints."""
    if isinstance(v, (list, tuple)):
        return {int(x) for x in v if str(x).isdigit()}
    return {int(m) for m in re.findall(r"\((\d+)/", v or "")}

def _valid_ip(ip):
    ip = (ip or "").strip()
    try:
        ipaddress.ip_address(ip)
        return ip not in ("0.0.0.0",)
    except Exception:
        return False

def _parse_date(s):
    for fmt in ("%Y-%m-%d", "%Y-%m", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.datetime.strptime(str(s)[:10], fmt).date()
        except Exception:
            continue
    return None

EXPOSURE_PORTS = {22, 23, 21, 3389, 5900}

def assemble(body):
    """
    Merge the three IT Glue / N-central sources into one device list plus a
    findings summary and a decommission (stale-AD) list. Join key is the IP:
    an N-central agent IP == a Network Glue discovered IP means the same device
    (so it's managed and we can promote the real hostname).

    Accepts:
      ncentral: [{name,class,ip,gw,dns,dhcp,mac,os}, ...]  (agent-managed)
      nonad:    [{ip,ports}, ...]                          (Network Glue scan)
      ad:       [{name,enabled,last}, ...]                 (Network Glue AD)
    Falls back to body['devices'] (flat, unmanaged-agnostic) when no ncentral.
    """
    ncentral = body.get("ncentral") or body.get("devices") or []
    nonad = body.get("nonad") or []
    ad = body.get("ad") or []
    stale_days = int(body.get("stale_days", 365))
    today = datetime.date.today()

    managed_ips, host_by_ip = set(), {}
    for d in ncentral:
        ip = (d.get("ip") or "").strip()
        if _valid_ip(ip):
            managed_ips.add(ip)
            host_by_ip.setdefault(ip, d.get("name"))

    devices, nonad_ips = [], set()
    for d in nonad:
        ip = (d.get("ip") or "").strip()
        if not _valid_ip(ip):
            continue
        nonad_ips.add(ip)
        devices.append({
            "name": host_by_ip.get(ip) or ip,
            "ip": ip,
            "class": "",
            "ports": d.get("ports"),
            "managed": ip in managed_ips,
        })
    for d in ncentral:
        ip = (d.get("ip") or "").strip()
        if not _valid_ip(ip) or ip in nonad_ips:
            continue
        dd = dict(d)
        dd["managed"] = True
        devices.append(dd)

    # Stale AD (enabled, no login for >= stale_days).
    # Make can't emit the domain backslash safely in JSON, so it sends the
    # AD name base64-encoded (name_b64); decode + strip the DOMAIN\ prefix here.
    stale = []
    for a in ad:
        en = a.get("enabled")
        if str(en).lower() != "true":
            continue
        dt = _parse_date(a.get("last"))
        if not (dt and (today - dt).days >= stale_days):
            continue
        raw = a.get("name")
        if not raw and a.get("name_b64"):
            try:
                raw = base64.b64decode(a["name_b64"]).decode("utf-8", "ignore")
            except Exception:
                raw = ""
        host = (raw or "").replace("/", "\\").split("\\")[-1]
        stale.append((host, str(a.get("last"))))
    stale.sort(key=lambda x: x[1])

    # Findings
    managed_ct = sum(1 for d in devices if d.get("managed"))
    disc = len(devices) - managed_ct
    exposed = sum(1 for d in nonad
                  if EXPOSURE_PORTS & parse_ports(d.get("ports")))
    findings = []
    if devices:
        findings.append("Coverage: %d managed / %d discovered-unmanaged"
                        % (managed_ct, disc))
    if stale:
        findings.append("Stale AD (>%dd): %d decommission candidates"
                        % (stale_days, len(stale)))
    if exposed:
        findings.append("Legacy exposure: %d hosts on SSH/Telnet/FTP/RDP" % exposed)
    return devices, findings, stale

def classify(name, cls, ports=None):
    n = (name or "").upper(); c = cls or ""
    if "FW" in n or "FIREWALL" in n: return "firewall"
    if "-AP" in n or "ACCESSPOINT" in n or "ACCESS POINT" in n or "-WAP" in n: return "accesspoint"
    if c == "Switch/Router" and "SW" in n: return "switch"
    if c == "Switch/Router": return "network"
    if c == "Printer" or "PRINT" in n or n.startswith("RNP"): return "printer"
    if c == "Storage": return "nas"
    if c.startswith("Servers"): return "hyperv" if "HV" in n else "server"
    if c.startswith("Laptop"): return "laptop"
    if c.startswith("Workstations"): return "endpoint"
    # No N-central class (discovered-only device): fingerprint by listening ports
    ps = ports or set()
    for port, role in PORT_ROLE:
        if port in ps:
            return role
    return "other"

def subnet_of(ip):
    try:
        return str(ipaddress.ip_network(ip + "/24", strict=False).network_address)[:-1] + "0/24"
    except Exception:
        return None

def nid(s): return "n_" + re.sub(r"\W", "_", s or "")
def esc(s): return (s or "").replace("\\", "").replace('"', "'")

def _dns_ips(devs):
    out = set()
    for d in devs:
        for x in (d.get("dns") or "").split(","):
            x = x.strip()
            if x and not x.startswith(PUBLIC_DNS):
                out.add(x)
    return out

def _seg_meta(ds):
    gws = [d.get("gw") for d in ds if d.get("gw")]
    gw = max(set(gws), key=gws.count) if gws else ""
    dnss = []
    for d in ds:
        for x in (d.get("dns") or "").split(","):
            x = x.strip()
            if x and not x.startswith(PUBLIC_DNS) and x not in dnss:
                dnss.append(x)
    dhcp = [x for x in {d.get("dhcp") for d in ds if d.get("dhcp")} if x]
    return gw, ", ".join(dnss[:3]), ", ".join(sorted(dhcp))

def _role_label(d, dns_ips):
    if d["role"] == "server" and d.get("ip") and d.get("ip") in dns_ips:
        return "DOMAIN CONTROLLER / DNS"
    return ROLE_NAMES.get(d["role"], "DEVICE")

def _node(d, dns_ips, suffix):
    col, fc = ROLE_COLOR.get(d["role"], ROLE_COLOR["other"])
    lines = [esc(d["name"])]
    if d.get("ip"): lines.append(d["ip"])
    lines.append(_role_label(d, dns_ips))
    os_s = esc((d.get("os") or "").replace("Microsoft ", ""))
    if os_s: lines.append(os_s)
    if d.get("mac"): lines.append(esc(d["mac"]))
    # Managed-vs-discovered is the coverage signal from IT Glue/N-central.
    # Discovered-but-unmanaged devices get a dashed border + a tag line.
    managed = d.get("managed", True)
    style = "filled,rounded" if managed else "filled,rounded,dashed"
    if not managed:
        lines.append("(DISCOVERED - unmanaged)")
    nidn = nid(d["name"] + "_" + suffix)
    return nidn, ('  %s [label="%s", fillcolor="%s", fontcolor="%s", style="%s"];'
                  % (nidn, "\\n".join(lines), col, fc, style))

def build_dot(client, devices, date, findings=None, stale=None):
    for d in devices:
        d["_ports"] = parse_ports(d.get("ports") or d.get("listening_ports"))
        d["role"] = classify(d.get("name"), d.get("class"), d["_ports"])
        d["ip"] = (d.get("ip") or "").strip()
        d["sub"] = subnet_of(d["ip"]) if d["ip"] else None
    dns_ips = _dns_ips(devices)
    fws = [d for d in devices if d["role"] == "firewall"]

    groups = {}
    for d in devices:
        groups.setdefault(d["sub"], []).append(d)
    real = {s: ds for s, ds in groups.items() if s}
    noip = groups.get(None, [])

    L = ["digraph G {",
         '  rankdir=TB; bgcolor="white"; fontname="Helvetica"; compound=true; nodesep=0.22; ranksep=0.7;',
         '  node [fontname="Helvetica", fontsize=9, style="filled,rounded", shape=box];',
         '  edge [color="#8a8a8a", arrowhead=none];',
         '  internet [label="INTERNET", shape=ellipse, fillcolor="#dddddd", fontcolor="#000000"];']

    # Findings panel (coverage gap / stale AD / exposure) from Network Glue data.
    if findings:
        ftext = "\\l".join(esc(x) for x in findings) + "\\l"
        L.append('  findings [label="NETWORK FINDINGS\\l%s", shape=note, '
                 'fillcolor="#fff6e6", fontcolor="#5a3b00", fontsize=10, '
                 'style="filled"];' % ftext)
        L.append('  {rank=source; internet; findings}')

    fw_anchor = "internet"
    for i, fw in enumerate(fws):
        fid = "fw%d" % i
        L.append('  %s [label="%s\\nFIREWALL / GATEWAY", fillcolor="#b3261e", fontcolor="white"];' % (fid, esc(fw["name"])))
        L.append('  internet -> %s;' % fid)
    if fws:
        fw_anchor = "fw0"

    order = {"firewall":0,"network":1,"switch":1,"accesspoint":2,"nas":3,"hyperv":4,"server":4,"printer":6,"endpoint":7,"laptop":8,"other":9}
    for idx, sub in enumerate(sorted(real.keys())):
        ds = real[sub]
        gw, dns, dhcp = _seg_meta(ds)
        meta = "Gateway %s" % (gw or "n/a")
        if dns: meta += "   DNS %s" % dns
        if dhcp: meta += "   DHCP %s" % dhcp
        anchor = "seg%d" % idx
        L.append('  subgraph cluster_seg%d {' % idx)
        L.append('    label="NETWORK SEGMENT  %s\\n%s"; style="rounded"; color="#1c7293"; penwidth=2; fontsize=12; fontcolor="#12315e";' % (sub, meta))
        L.append('    %s [label="%s\\nSEGMENT GATEWAY", shape=box, style="filled", fillcolor="#1c7293", fontcolor="white"];' % (anchor, gw or sub))
        devs = [d for d in sorted(ds, key=lambda x: order.get(x["role"], 9)) if d["role"] != "firewall"]
        ids = []
        for d in devs:
            nidn, line = _node(d, dns_ips, "s%d" % idx)
            L.append("  " + line)
            ids.append(nidn)
        rows = [ids[i:i + 6] for i in range(0, len(ids), 6)]
        for row in rows:
            L.append("    {rank=same; " + " ".join(row) + "}")
            for j in range(len(row) - 1):
                L.append('    %s -> %s [style=invis];' % (row[j], row[j + 1]))
        for r in range(len(rows) - 1):
            L.append('    %s -> %s [style=invis];' % (rows[r][0], rows[r + 1][0]))
        if rows:
            L.append('    %s -> %s;' % (anchor, rows[0][0]))
        L.append('  }')
        L.append('  %s -> %s [lhead=cluster_seg%d, color="#b3261e", penwidth=1.4];' % (fw_anchor, anchor, idx))

    if stale:
        L.append('  subgraph cluster_stale {')
        L.append('    label="DECOMMISSION CANDIDATES  -  enabled in AD, no login >1yr"; '
                 'style="dashed"; color="#b3261e"; fontsize=11; fontcolor="#b3261e";')
        sids = []
        for i, (nm, last) in enumerate(stale):
            sid = "stale%d" % i
            L.append('    %s [label="%s\\n%s", fillcolor="#f3d9d6", fontcolor="#7a1f2b", shape=box];'
                     % (sid, esc(nm), esc(last)))
            sids.append(sid)
        srows = [sids[i:i + 8] for i in range(0, len(sids), 8)]
        for row in srows:
            L.append("    {rank=same; " + " ".join(row) + "}")
            for j in range(len(row) - 1):
                L.append('    %s -> %s [style=invis];' % (row[j], row[j + 1]))
        for r in range(len(srows) - 1):
            L.append('    %s -> %s [style=invis];' % (srows[r][0], srows[r + 1][0]))
        L.append('  }')

    if noip:
        L.append('  subgraph cluster_noip {')
        L.append('    label="DISCOVERED DEVICES WITHOUT AN IP (SNMP / agent-only)"; style="dashed"; color="#999999"; fontsize=11;')
        nids = []
        for d in noip:
            nidn, line = _node(d, dns_ips, "noip")
            L.append("  " + line)
            nids.append(nidn)
        nrows = [nids[i:i + 8] for i in range(0, len(nids), 8)]
        for row in nrows:
            L.append("    {rank=same; " + " ".join(row) + "}")
            for j in range(len(row) - 1):
                L.append('    %s -> %s [style=invis];' % (row[j], row[j + 1]))
        for r in range(len(nrows) - 1):
            L.append('    %s -> %s [style=invis];' % (nrows[r][0], nrows[r + 1][0]))
        L.append('  }')

    L.append("}")
    return "\n".join(L)

def brand(png_bytes, logo_url, client, date):
    diag = Image.open(io.BytesIO(png_bytes)).convert("RGBA")
    W, H = diag.size
    MAXW = 11000
    if W > MAXW:
        diag = diag.resize((MAXW, int(H * (MAXW / W))), Image.LANCZOS)
        W, H = diag.size
    head, foot = 100, 62
    f_foot = ImageFont.truetype(FONT_BOLD, 15)
    f_t1 = ImageFont.truetype(FONT_BOLD, 26)
    f_t2 = ImageFont.truetype(FONT_BOLD, 16)
    meas = ImageDraw.Draw(Image.new("RGB", (1, 1)))
    foot_tw = meas.textlength(CONF, font=f_foot)
    logo = None
    if logo_url:
        try:
            raw = urllib.request.urlopen(logo_url, timeout=15).read()
            logo = Image.open(io.BytesIO(raw)).convert("RGBA"); logo.thumbnail((84, 84))
        except Exception as e:
            app.logger.warning("logo fetch failed: %s", e)
    logo_w = logo.width if logo else 84
    title_x = 28 + logo_w + 28
    t1 = "%s  -  Network Map (logical)" % client
    t2 = "auto-generated from N-central, %s" % date
    t1w = meas.textlength(t1, font=f_t1); t2w = meas.textlength(t2, font=f_t2)
    CW = int(max(W, foot_tw + 160, title_x + max(t1w, t2w) + 40))
    canvas = Image.new("RGBA", (CW, head + H + foot), (255, 255, 255, 255))
    canvas.paste(diag, ((CW - W) // 2, head))
    if logo:
        canvas.paste(logo, (28, (head - logo.height) // 2), logo)
    d = ImageDraw.Draw(canvas)
    d.text((title_x, head / 2 - 30), t1, font=f_t1, fill=NAVY)
    d.text((title_x, head / 2 + 6), t2, font=f_t2, fill=(90, 90, 90))
    d.rectangle([0, head + H, CW, head + H + foot], fill=NAVY)
    d.text((CW / 2 - foot_tw / 2, head + H + foot / 2 - 10), CONF, font=f_foot, fill=(255, 255, 255))
    out = io.BytesIO(); canvas.convert("RGB").save(out, "PNG"); out.seek(0)
    return out

# ---------------------------------------------------------------------------
# Rich-text inventory for the IT Glue "Map Notes" field (kind: Textbox/HTML)
# ---------------------------------------------------------------------------
def _host_octet(ip):
    try:
        return int(str(ip).split(",")[0].strip().split(".")[-1])
    except Exception:
        return 999

def _netkey(sub):
    try:
        return (0, int(ipaddress.ip_network(sub).network_address))
    except Exception:
        return (1, sub)

def _sublabel(devs):
    roles = " ".join(d["role"] for d in devs).upper()
    names = " ".join(d["name"].upper() for d in devs)
    if "DOMAIN CONTROLLER" in roles or "ORACLE" in roles or "SQL" in roles:
        return "server / datacenter VLAN"
    if "LINUX" in roles:
        return "Linux / appliance VLAN"
    if names.count("LAPTOP") and len(devs) <= 6:
        return "VPN / remote pool"
    return "LAN segment"

def _h(s):
    return (str(s) if s is not None else "").replace("&", "&amp;").replace(
        "<", "&lt;").replace(">", "&gt;")

def build_notes_html(client, date, devices, findings, stale):
    rows = []
    for d in devices:
        ip = (d.get("ip") or "").strip()
        ports = parse_ports(d.get("ports") or d.get("listening_ports"))
        role = ROLE_NAMES.get(classify(d.get("name"), d.get("class"), ports), "DEVICE")
        rows.append({"name": d.get("name") or ip or "(unnamed)", "ip": ip or "-",
                     "role": role, "managed": d.get("managed", True),
                     "sub": subnet_of(ip) if ip else "no-IP", "octet": _host_octet(ip)})
    groups = OrderedDict()
    for r in sorted(rows, key=lambda r: (_netkey(r["sub"] or "z"), r["octet"])):
        groups.setdefault(r["sub"] or "no-IP", []).append(r)

    ts = 'border="1" cellpadding="6" cellspacing="0" style="border-collapse:collapse;width:100%"'
    out = ['<h2>%s &mdash; network inventory</h2>' % _h(client),
           '<p><em>Source: N-central + Network Glue &middot; %s</em></p>' % _h(date)]
    if findings:
        out.append("<p><strong>Findings</strong></p><ul>")
        out += ["<li>%s</li>" % _h(f) for f in findings]
        out.append("</ul>")
    for sub, devs in groups.items():
        managed = sum(1 for x in devs if x["managed"])
        out.append('<h3>%s &mdash; %s &middot; %d devices &middot; %d managed</h3>'
                   % (_h(sub), _h(_sublabel(devs)), len(devs), managed))
        out.append('<table %s><tr><th align="left">Device</th><th align="left">IP</th>'
                   '<th align="left">Role</th><th align="left">Managed</th></tr>' % ts)
        for x in devs:
            tag = "Agent" if x["managed"] else "Discovered"
            out.append("<tr><td>%s</td><td>%s</td><td>%s</td><td>%s</td></tr>"
                       % (_h(x["name"]), _h(x["ip"]), _h(x["role"]), tag))
        out.append("</table>")
    if stale:
        out.append('<h3>Decommission candidates &mdash; enabled in AD, no login &gt;1&nbsp;yr</h3>')
        out.append('<table %s><tr><th align="left">Computer</th><th align="left">Last login</th></tr>' % ts)
        for nm, last in stale:
            out.append("<tr><td>%s</td><td>%s</td></tr>" % (_h(nm), _h(last)))
        out.append("</table>")
    return "".join(out)

@app.route("/notes", methods=["POST"])
def notes():
    key = os.environ.get("RENDER_KEY")
    if key and request.headers.get("X-Api-Key") != key:
        return jsonify(error="unauthorized"), 401
    body = request.get_json(force=True)
    client = body.get("client", "Client")
    date = body.get("date") or datetime.datetime.now().strftime("%Y-%m-%d")
    devices, findings, stale = assemble(body)
    html = build_notes_html(client, date, devices, findings, stale)
    # Return JSON-string-escaped HTML (no surrounding quotes) so the Make
    # scenario can drop it straight into a JSON PATCH body between quotes.
    return app.response_class(json.dumps(html)[1:-1], mimetype="text/plain")

@app.route("/healthz")
def healthz(): return "ok", 200

@app.route("/render", methods=["POST"])
def render():
    key = os.environ.get("RENDER_KEY")
    if key and request.headers.get("X-Api-Key") != key:
        return jsonify(error="unauthorized"), 401
    body = request.get_json(force=True)
    client = body.get("client", "Client")
    date = body.get("date") or datetime.datetime.now().strftime("%Y-%m-%d")
    # Enriched mode: merge ncentral + Network Glue nonad + ad server-side.
    devices, findings, stale = assemble(body)
    if not devices:
        return jsonify(error="no devices provided"), 400
    if body.get("findings"):          # allow caller-supplied override lines
        findings = list(body["findings"])
    dot = build_dot(client, devices, date, findings, stale)
    p = subprocess.run(["dot", "-Tpng", "-Gdpi=110"], input=dot.encode(),
                       stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    if p.returncode != 0:
        return jsonify(error="graphviz failed", detail=p.stderr.decode()[:500]), 500
    out = brand(p.stdout, body.get("logo_url"), client, date)
    return send_file(out, mimetype="image/png",
                     download_name=re.sub(r"\W+", "_", client) + "_network_map.png")

# ===================== TrustPoint Patch & Vulnerability report  (v3.2) =====================
# Added alongside the network-map service. POST /patch-report (JSON):
#   { "customer": "...",
#     "summary":    [ {hostName, os, scanTimeUtc, critical, important, moderate, low, appCount, lastPatched}, ... ],
#     "detections": [ {hostName, os, itemType, ref, title, severity, releaseDate, available, scanTimeUtc}, ... ] }
# -> 200 xlsx  (or JSON {filename, dataB64, stats} when called with ?b64=1). X-Api-Key (RENDER_KEY) like /render.
#
# v3.2 adds:
#   * "Last patched" per device (from the device's own Windows Update install history)
#   * Overdue split: a Windows/MS patch whose releaseDate is older than the 14-day hold
#     and still missing = OVERDUE ("patching not working"); newer = still in-hold (expected)
#   * Overdue KPI + alert banner on the Summary sheet
#   * stats returned in the JSON response so the orchestrator can raise an internal alert
import tempfile, base64
import datetime as _dt
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as _plt
from openpyxl import Workbook as _WB
from openpyxl.styles import Font as _F, PatternFill as _PF, Alignment as _AL, Border as _BD, Side as _SD
from openpyxl.utils import get_column_letter as _gcl
from openpyxl.worksheet.properties import PageSetupProperties as _PSP
from openpyxl.drawing.image import Image as _XLImage

_NAVY="1F2A44"; _BLUE="2E5BFF"; _SLATE="5B6B8C"
_CRIT="C0392B"; _IMP="E67E22"; _MOD="F1C40F"; _LOW="27AE60"; _APPC="2E86C1"; _OVER="B03A2E"
_LIGHT="EEF2FB"; _WHITE="FFFFFF"; _AMBER="FBEEDC"; _FONT="Arial"
_thin=_SD(style="thin",color="D5DBE8"); _bd=_BD(left=_thin,right=_thin,top=_thin,bottom=_thin)
_sevc={"Critical":_CRIT,"Important":_IMP,"Moderate":_MOD,"Low":_LOW,"App updates":_APPC}

HOLD_DAYS = 14   # TrustPoint patch policy: 14-day approval hold. Past this + still missing = overdue.

def _hx(c): return "#"+c
def _i(v):
    try: return int(float(v))
    except Exception: return 0

def _patch_parse_date(s):
    """Parse an ISO date/datetime like '2026-08-13' or '2026-08-13T04:00:00Z'; return date or None."""
    if not s: return None
    s=str(s).strip()
    for fmt in ("%Y-%m-%d","%Y-%m-%dT%H:%M:%SZ","%Y-%m-%dT%H:%M:%S","%Y-%m-%d %H:%M:%S"):
        try: return _dt.datetime.strptime(s,fmt).date()
        except Exception: pass
    try: return _dt.datetime.fromisoformat(s.replace("Z","")).date()
    except Exception: return None

def _is_overdue(item_type, release_date, today):
    """A Windows/Microsoft patch released more than HOLD_DAYS ago and still missing is overdue."""
    if not item_type or "Windows" not in item_type and "Microsoft" not in item_type:
        return (False, None)
    d=_patch_parse_date(release_date)
    if not d: return (False, None)
    age=(today-d).days
    return (age>HOLD_DAYS, age)

def _last_patched_display(iso, today):
    d=_patch_parse_date(iso)
    if not d: return ("unknown", None)
    days=(today-d).days
    return (d.strftime("%Y-%m-%d")+f"  ({days}d ago)", days)

def _rdonut(sev,total,path):
    labels=[k for k,v in sev]; vals=[v for k,v in sev]; colors=[_hx(_sevc[k]) for k in labels]
    fig,ax=_plt.subplots(figsize=(6.4,3.3),dpi=110)
    w,_=ax.pie(vals if any(vals) else [1],colors=colors if any(vals) else ["#DDDDDD"],startangle=90,counterclock=False,wedgeprops=dict(width=0.42,edgecolor="white",linewidth=1.5))
    ax.set(aspect="equal")
    ax.text(0,0,str(total)+"\nitems",ha="center",va="center",fontsize=15,fontweight="bold",color=_hx(_NAVY))
    leg=[labels[i]+"   "+str(vals[i])+"  ("+(f"{(vals[i]/total*100):.1f}" if total else "0.0")+"%)" for i in range(len(labels))]
    ax.legend(w[:len(labels)],leg,loc="center left",bbox_to_anchor=(1.0,0.5),frameon=False,fontsize=10.5)
    _plt.subplots_adjust(left=0.02,right=0.62,top=0.98,bottom=0.02)
    fig.savefig(path,dpi=110,facecolor="white"); _plt.close(fig)

def _rbars(topdev,path):
    names=[d[0] for d in topdev][::-1]; totals=[d[2]+d[3]+d[4]+d[5]+d[6] for d in topdev][::-1]; n=len(names)
    fig,ax=_plt.subplots(figsize=(9.2,max(1.6,0.46*n+0.7)),dpi=110)
    ax.barh(range(n),totals,color=_hx(_BLUE),height=0.62)
    ax.set_yticks(range(n)); ax.set_yticklabels(names,fontsize=10.5,color=_hx(_NAVY)); ax.tick_params(axis="y",length=0)
    ax.tick_params(axis="x",labelsize=9,colors=_hx(_SLATE))
    for s in ("top","right","left"): ax.spines[s].set_visible(False)
    ax.spines["bottom"].set_color(_hx("D5DBE8")); mx=max(totals) if totals else 1; ax.set_xlim(0,mx*1.12)
    for i,v in enumerate(totals): ax.text(v+mx*0.015,i,str(v),va="center",ha="left",fontsize=9.5,color=_hx(_NAVY))
    _plt.subplots_adjust(left=0.28,right=0.98,top=0.98,bottom=0.14)
    fig.savefig(path,dpi=110,facecolor="white"); _plt.close(fig)

def build_patch_report(customer, summary, detections):
    today=_dt.date.today()
    report_date=_dt.datetime.now().strftime("%B %d, %Y")

    # ---- overdue per host (from detections) ----
    overdue_by_host={}
    overdue_total=0; oldest_overdue_days=0
    det_overdue_flags=[]   # parallel to detections order below
    for d in detections:
        od,age=_is_overdue(d.get("itemType",""), d.get("releaseDate",""), today)
        det_overdue_flags.append((od,age))
        if od:
            h=d.get("hostName","")
            overdue_by_host[h]=overdue_by_host.get(h,0)+1
            overdue_total+=1
            if age and age>oldest_overdue_days: oldest_overdue_days=age

    # ---- devices (summary) with last-patched + overdue count ----
    devices=[]
    lp_known_days=[]; lp_unknown=0
    for s in summary:
        host=s.get("hostName","")
        lp_disp,lp_days=_last_patched_display(s.get("lastPatched",""), today)
        if lp_days is None: lp_unknown+=1
        else: lp_known_days.append((host,lp_days))
        devices.append([host, s.get("os",""), _i(s.get("critical")), _i(s.get("important")),
                        _i(s.get("moderate")), _i(s.get("low")), _i(s.get("appCount")),
                        overdue_by_host.get(host,0), lp_disp, s.get("scanTimeUtc","")])

    details=[]
    for idx,d in enumerate(detections):
        od,age=det_overdue_flags[idx]
        details.append([d.get("hostName",""), d.get("os",""), d.get("itemType",""), d.get("ref",""),
                        d.get("title",""), d.get("severity","Unspecified") or "Unspecified",
                        d.get("releaseDate","") or "", ("YES" if od else ""),
                        d.get("available","") or "", d.get("scanTimeUtc","")])

    tc=sum(d[2] for d in devices); ti=sum(d[3] for d in devices); tm=sum(d[4] for d in devices)
    tl=sum(d[5] for d in devices); ta=sum(d[6] for d in devices); tall=tc+ti+tm+tl+ta
    devices_overdue=sum(1 for h,c in overdue_by_host.items() if c>0)
    ranked=sorted(devices,key=lambda d:(d[7]*10000)+(d[2]+d[3]+d[4]+d[5]+d[6]),reverse=True)  # overdue first
    sev=[("Critical",tc),("Important",ti),("Moderate",tm),("Low",tl),("App updates",ta)]
    topdev=ranked[:12]

    stats={
        "overdueTotal":overdue_total, "devicesOverdue":devices_overdue,
        "oldestOverdueDays":oldest_overdue_days, "devicesScanned":len(devices),
        "totalMissing":tall, "critical":tc, "important":ti,
        "lastPatchedUnknown":lp_unknown,
        "oldestLastPatchedDays":(max((d for _,d in lp_known_days), default=0)),
        "oldestLastPatchedHost":(max(lp_known_days, key=lambda x:x[1])[0] if lp_known_days else ""),
    }

    dp=tempfile.NamedTemporaryFile(suffix=".png",delete=False).name
    bp=tempfile.NamedTemporaryFile(suffix=".png",delete=False).name
    _rdonut(sev,tall,dp); _rbars(topdev if topdev else [["(no devices)","",0,0,0,0,0,0,"",""]],bp)

    wb=_WB(); ws=wb.active; ws.title="Summary"; ws.sheet_view.showGridLines=False
    for c in range(1,16): ws.column_dimensions[_gcl(c)].width=11
    ws.merge_cells("A1:O2"); t=ws["A1"]; t.value=customer+"  -  Patch & Vulnerability Report"
    t.font=_F(name=_FONT,size=18,bold=True,color=_WHITE); t.alignment=_AL(horizontal="left",vertical="center",indent=1)
    for row in ws["A1:O2"]:
        for cell in row: cell.fill=_PF("solid",fgColor=_NAVY)
    ws.merge_cells("A3:O3"); s=ws["A3"]
    s.value="Week of "+report_date+"    -    Source: on-device Windows Update + winget scan (machine-sourced, accurate)"
    s.font=_F(name=_FONT,size=10,italic=True,color=_SLATE); s.alignment=_AL(horizontal="left",indent=1)

    kpis=[("Devices scanned",len(devices),_BLUE),("Total missing",tall,_NAVY),("Critical",tc,_CRIT),
          ("Important",ti,_IMP),("App updates",ta,_APPC),("Overdue (>%dd)"%HOLD_DAYS,overdue_total,_OVER)]
    col=1
    for label,val,color in kpis:
        ws.merge_cells(start_row=5,start_column=col,end_row=5,end_column=col+1)
        ws.merge_cells(start_row=6,start_column=col,end_row=7,end_column=col+1)
        lc=ws.cell(row=5,column=col,value=label); lc.font=_F(name=_FONT,size=9,bold=True,color=_WHITE); lc.alignment=_AL(horizontal="center")
        vc=ws.cell(row=6,column=col,value=val); vc.font=_F(name=_FONT,size=22,bold=True,color=_WHITE); vc.alignment=_AL(horizontal="center",vertical="center")
        for r in (5,6,7):
            for cc in (col,col+1): ws.cell(row=r,column=cc).fill=_PF("solid",fgColor=color)
        col+=2

    # ---- overdue / all-clear banner (row 8) ----
    ws.merge_cells("A8:O8"); bn=ws["A8"]
    if overdue_total>0:
        bn.value=("  ATTENTION: %d approved patch(es) on %d device(s) are past the %d-day hold and still missing "
                  "- patching may not be completing. See 'By Device' (Overdue column) and 'Detections' (Overdue = YES).") % (overdue_total,devices_overdue,HOLD_DAYS)
        bn.font=_F(name=_FONT,size=10,bold=True,color=_OVER); bn.fill=_PF("solid",fgColor=_AMBER)
    else:
        bn.value="  All missing patches are within the %d-day approval hold - nothing overdue. Patching is current." % HOLD_DAYS
        bn.font=_F(name=_FONT,size=10,bold=True,color=_LOW); bn.fill=_PF("solid",fgColor=_LIGHT)
    bn.alignment=_AL(horizontal="left",vertical="center")

    ws.merge_cells("A10:E10"); ws["A10"].value="Missing items by severity"; ws["A10"].font=_F(name=_FONT,size=12,bold=True,color=_NAVY)
    ws.merge_cells("H10:J10"); ws["H10"].value="Breakdown"; ws["H10"].font=_F(name=_FONT,size=12,bold=True,color=_NAVY)
    im=_XLImage(dp); im.width=470; im.height=242; ws.add_image(im,"A11")
    for i,h in enumerate(["Severity","Count","Share"],start=8):
        cell=ws.cell(row=11,column=i,value=h); cell.font=_F(name=_FONT,size=10,bold=True,color=_WHITE)
        cell.fill=_PF("solid",fgColor=_NAVY); cell.alignment=_AL(horizontal="center"); cell.border=_bd
    r=12
    for k,v in sev:
        kc=ws.cell(row=r,column=8,value=k); kc.font=_F(name=_FONT,size=10,bold=True,color=_WHITE)
        kc.fill=_PF("solid",fgColor=_sevc[k]); kc.alignment=_AL(horizontal="left",indent=1); kc.border=_bd
        vc=ws.cell(row=r,column=9,value=v); vc.font=_F(name=_FONT,size=10,color=_NAVY); vc.alignment=_AL(horizontal="center"); vc.border=_bd
        pc=ws.cell(row=r,column=10,value="=I"+str(r)+"/I17"); pc.number_format="0.0%"
        pc.font=_F(name=_FONT,size=10,color=_NAVY); pc.alignment=_AL(horizontal="center"); pc.border=_bd
        r+=1
    tcc=ws.cell(row=17,column=8,value="Total"); tcc.font=_F(name=_FONT,size=10,bold=True,color=_WHITE)
    tcc.fill=_PF("solid",fgColor=_SLATE); tcc.alignment=_AL(horizontal="left",indent=1); tcc.border=_bd
    tv=ws.cell(row=17,column=9,value="=SUM(I12:I16)"); tv.font=_F(name=_FONT,size=10,bold=True,color=_WHITE)
    tv.fill=_PF("solid",fgColor=_SLATE); tv.alignment=_AL(horizontal="center"); tv.border=_bd
    tp=ws.cell(row=17,column=10,value="=I17/I17"); tp.number_format="0.0%"; tp.font=_F(name=_FONT,size=10,bold=True,color=_WHITE)
    tp.fill=_PF("solid",fgColor=_SLATE); tp.alignment=_AL(horizontal="center"); tp.border=_bd

    ws.merge_cells("A29:O29"); ws["A29"].value="Top devices (overdue first, then by items missing)"; ws["A29"].font=_F(name=_FONT,size=12,bold=True,color=_NAVY)
    bi=_XLImage(bp); sc=760.0/bi.width; bi.width=760; bi.height=int(bi.height*sc); ws.add_image(bi,"A30")

    # ---- By Device sheet ----
    wc=wb.create_sheet("By Device"); wc.sheet_view.showGridLines=False
    headers=["Device","Operating System","Critical","Important","Moderate","Low","App updates","Total","Overdue (>%dd)"%HOLD_DAYS,"Last patched","Last scanned"]
    widths=[24,26,10,11,11,8,12,10,13,20,20]
    for i,(h,w) in enumerate(zip(headers,widths),start=1):
        cell=wc.cell(row=1,column=i,value=h); cell.font=_F(name=_FONT,size=10,bold=True,color=_WHITE)
        cell.fill=_PF("solid",fgColor=(_OVER if i==9 else _NAVY)); cell.alignment=_AL(horizontal="center" if 3<=i<=9 else "left",vertical="center"); cell.border=_bd
        wc.column_dimensions[_gcl(i)].width=w
    wc.row_dimensions[1].height=22
    ncols=len(headers)
    for rr,d in enumerate(ranked,start=2):
        total=d[2]+d[3]+d[4]+d[5]+d[6]
        vals=[d[0],d[1],d[2],d[3],d[4],d[5],d[6],total,d[7],d[8],d[9]]  # d[7]=overdue, d[8]=last patched, d[9]=last scanned
        for i,v in enumerate(vals,start=1):
            cell=wc.cell(row=rr,column=i,value=v); cell.font=_F(name=_FONT,size=10,color=_NAVY)
            cell.alignment=_AL(horizontal="center" if 3<=i<=9 else "left"); cell.border=_bd
            if rr%2==0: cell.fill=_PF("solid",fgColor=_LIGHT)
        wc.cell(row=rr,column=3).font=_F(name=_FONT,size=10,bold=True,color=_CRIT)
        wc.cell(row=rr,column=4).font=_F(name=_FONT,size=10,bold=True,color=_IMP)
        ov=wc.cell(row=rr,column=9)
        if d[7]>0:
            ov.font=_F(name=_FONT,size=10,bold=True,color=_WHITE); ov.fill=_PF("solid",fgColor=_OVER)
        if str(d[8]).startswith("unknown"):
            wc.cell(row=rr,column=10).font=_F(name=_FONT,size=10,italic=True,color=_SLATE)
    tr=len(ranked)+2
    for i in range(1,ncols+1):
        cell=wc.cell(row=tr,column=i); cell.fill=_PF("solid",fgColor=_SLATE); cell.font=_F(name=_FONT,bold=True,color=_WHITE)
        cell.alignment=_AL(horizontal="center" if 3<=i<=9 else "left"); cell.border=_bd
    wc.cell(row=tr,column=1,value="TOTAL")
    for i,cl in enumerate(["C","D","E","F","G","H","I"],start=3): wc.cell(row=tr,column=i,value="=SUM("+cl+"2:"+cl+str(tr-1)+")")
    wc.freeze_panes="A2"

    # ---- Detections sheet ----
    wd=wb.create_sheet("Detections"); wd.sheet_view.showGridLines=False
    dh=["Device","OS","Type","KB / App ID","Title","Severity","Released","Overdue?","Fixed / Available","Last scanned"]; dw=[22,18,17,26,40,12,13,10,20,18]
    for i,(h,w) in enumerate(zip(dh,dw),start=1):
        cell=wd.cell(row=1,column=i,value=h); cell.font=_F(name=_FONT,size=10,bold=True,color=_WHITE)
        cell.fill=_PF("solid",fgColor=_NAVY); cell.alignment=_AL(horizontal="left",vertical="center"); cell.border=_bd
        wd.column_dimensions[_gcl(i)].width=w
    wd.row_dimensions[1].height=22
    # overdue first, then by severity weight
    _sevorder={"Critical":0,"Important":1,"Moderate":2,"Low":3,"Unspecified":4}
    details_sorted=sorted(details,key=lambda r:(0 if r[7]=="YES" else 1, _sevorder.get(r[5],9)))
    for rr,row in enumerate(details_sorted,start=2):
        for i,v in enumerate(row,start=1):
            cell=wd.cell(row=rr,column=i,value=v); cell.font=_F(name=_FONT,size=9,color=_NAVY)
            cell.border=_bd; cell.alignment=_AL(horizontal="left",vertical="center")
            if rr%2==0: cell.fill=_PF("solid",fgColor=_LIGHT)
        scc=wd.cell(row=rr,column=6); scc.font=_F(name=_FONT,size=9,bold=True,color=_WHITE)
        scc.fill=_PF("solid",fgColor=_sevc.get(row[5],_SLATE)); scc.alignment=_AL(horizontal="center",vertical="center")
        if row[7]=="YES":
            oc=wd.cell(row=rr,column=8); oc.font=_F(name=_FONT,size=9,bold=True,color=_WHITE)
            oc.fill=_PF("solid",fgColor=_OVER); oc.alignment=_AL(horizontal="center",vertical="center")
    wd.freeze_panes="A2"

    for sh in wb.worksheets:
        sh.page_setup.orientation="landscape"; sh.page_setup.fitToWidth=1; sh.page_setup.fitToHeight=0
        sh.sheet_properties.pageSetUpPr=_PSP(fitToPage=True)
        sh.page_margins.left=0.3; sh.page_margins.right=0.3; sh.page_margins.top=0.4; sh.page_margins.bottom=0.4
    ws.print_area="A1:O"+str(31+len(topdev)+14)
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    try: os.remove(dp); os.remove(bp)
    except Exception: pass
    return buf, stats

@app.route("/patch-report", methods=["POST"])
def patch_report():
    key = os.environ.get("RENDER_KEY")
    if key and request.headers.get("X-Api-Key") != key:
        return jsonify(error="unauthorized"), 401
    body = request.get_json(force=True)
    customer = body.get("customer", "Client")
    summary = body.get("summary", [])
    detections = body.get("detections", [])
    if not summary:
        return jsonify(error="no devices provided"), 400
    buf, stats = build_patch_report(customer, summary, detections)
    fname = re.sub(r"\W+", "_", customer) + "_patch_report.xlsx"
    if request.args.get("b64"):
        return jsonify(filename=fname, dataB64=base64.b64encode(buf.getvalue()).decode(), stats=stats)
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", download_name=fname)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 8080)))
